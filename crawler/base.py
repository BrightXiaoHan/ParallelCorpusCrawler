import pickle
import logging

from openpyxl import load_workbook, Workbook
from typing import Iterable, List
from tqdm import tqdm


class BaseCrawler(object):

    def __init__(self,
                 src_kw: List[str],
                 tgt_kw: List[str]) -> None:

        self.src_kw = src_kw
        self.tgt_kw = tgt_kw

        self.error_list = {
            "src": [],
            "tgt": []
        }
        self.corpus = []

    @classmethod
    def from_excel(cls,
                   excel_path: str,
                   src_column: int = 0,
                   tgt_column: int = 1,
                   skip_first_column: bool = True,
                   **kwargs):
        src_words, tgt_words = [], []
        sheet = load_workbook(filename=excel_path, read_only=True).active

        for i, row in tqdm(enumerate(sheet.iter_rows())):
            if not i and skip_first_column:
                continue
            src = row[src_column].value
            tgt = row[tgt_column].value

            if src and tgt:
                src_words.append(str(src))
                tgt_words.append(str(tgt))

        return cls(src_words, tgt_words, **kwargs)

    @classmethod
    def from_checkpoint(cls, checkpoint_path: str):
        f = open(checkpoint_path, 'rb')
        obj = pickle.load(f)
        assert isinstance(obj, cls)
        f.close()

    def crawl(self):
        for i, seeds in tqdm(enumerate(zip(self.src_kw, self.tgt_kw))):
            for seed, tp in zip(seeds, ["src", "tgt"]):
                try:
                    for pair in self._crawl_single(seed, type=tp):
                        pair.extend(seeds)
                        self.corpus.append(pair)
                except Exception as e:
                    logging.log(
                        level=logging.WARNING, msg="Crawl with seed %s error. Error message: %s" % (seed, e))
                    self.error_list[tp].append(seed)

    def save(self, file_path: str):
        wb = Workbook()
        sheet = wb.active

        sheet.cell(row=1, column=1, value="Source Sentence")
        sheet.cell(row=1, column=2, value="Target Sentence")
        sheet.cell(row=1, column=3, value="Seed Source")
        sheet.cell(row=1, column=4, value="Seed Target")

        for i, (src, tgt, src_seed, tgt_seed) in enumerate(self.corpus):
            i += 2
            sheet.cell(row=i, column=1, value=src)
            sheet.cell(row=i, column=2, value=tgt)
            sheet.cell(row=i, column=3, value=src_seed)
            sheet.cell(row=i, column=4, value=tgt_seed)

        wb.save(file_path)

    def _crawl_single(self, seed: str, type: str = "src") -> Iterable[List[str]]:
        raise NotImplementedError
